#!/usr/bin/env python3
"""Build separate Excel reports for the two experimental studies.

Study A: A_B1_VS_B2
    Main protocol evaluation, no forced loss.

Study B: B_V1_V2_FIRST_LOSS
    Targeted validation of V1 vs V2 with a forced first-fragment loss.

The script reads the global CSV files in ``results/`` and writes one workbook
per campaign and ML-KEM level under ``results/workbooks/<campaign>/``.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

KNOWN_SECURITY_LEVELS = [512, 768, 1024]
KNOWN_CAMPAIGNS = ["A_B1_VS_B2", "B_V1_V2_FIRST_LOSS"]

INPUT_FILES = {
    "summary": "summary.csv",
    "all_runs": "all_runs.csv",
    "failed_runs": "failed_runs.csv",
    "log_manifest": "log_manifest.csv",
    "validation_issues": "validation_issues.csv",
}

SHEET_TITLES = {
    "summary": "Résumé",
    "all_runs": "Runs (tous)",
    "failed_runs": "Runs échoués",
    "log_manifest": "Manifest logs",
    "validation_issues": "Anomalies",
    "console_log": "Console (extrait)",
}

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=10)
MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 8

RUN_ID_KEM_RE = re.compile(r"^(?:[AB]-)?kem(\d+)[_-]")


def load_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        print(f"  [!] Fichier manquant, onglet ignoré : {path.name}", file=sys.stderr)
        return pd.DataFrame()
    return pd.read_csv(path)


def load_console_logs(input_dir: Path) -> list[str]:
    """Merge all campaign console .log files found directly in results/."""
    lines: list[str] = []
    for path in sorted(input_dir.glob("*.log")):
        try:
            lines.extend(path.read_text(encoding="utf-8", errors="replace").splitlines())
        except OSError:
            print(f"  [!] Log illisible : {path.name}", file=sys.stderr)
    return lines


def infer_kem_level_from_run_id(run_id: str) -> int | None:
    match = RUN_ID_KEM_RE.match(str(run_id))
    return int(match.group(1)) if match else None


def infer_campaign_from_run_id(run_id: str) -> str | None:
    run_id = str(run_id)
    if run_id.startswith("A-kem"):
        return "A_B1_VS_B2"
    if run_id.startswith("B-kem"):
        return "B_V1_V2_FIRST_LOSS"
    return None


def normalize_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Add campaign/kem_level when they can be inferred from run_id."""
    if df.empty:
        return df
    df = df.copy()
    if "campaign" not in df.columns and "run_id" in df.columns:
        df["campaign"] = df["run_id"].map(infer_campaign_from_run_id)
    if "kem_level" not in df.columns and "run_id" in df.columns:
        df["kem_level"] = df["run_id"].map(infer_kem_level_from_run_id)
    return df


def autosize_and_style(ws: Worksheet, n_rows: int, n_cols: int) -> None:
    if n_cols == 0:
        return
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=False)
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1, max_col=n_cols):
        for cell in row:
            cell.font = BODY_FONT
    ws.freeze_panes = "A2"
    if n_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        header_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        sample_lens = [
            len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            for row_idx in range(2, min(n_rows + 1, 200) + 1)
        ]
        width = max([header_len] + sample_lens + [MIN_COL_WIDTH])
        ws.column_dimensions[letter].width = min(width + 2, MAX_COL_WIDTH)


def write_dataframe_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=title[:31])
    if df.empty:
        ws.append(["(aucune donnée)"])
        ws["A1"].font = BODY_FONT
        return
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    autosize_and_style(ws, len(df), len(df.columns))


def write_console_sheet(wb: Workbook, title: str, lines: list[str]) -> None:
    ws = wb.create_sheet(title=title[:31])
    ws.append(["ligne_log"])
    if not lines:
        ws.append(["(aucune ligne)"])
    else:
        for line in lines:
            ws.append([line])
    autosize_and_style(ws, max(len(lines), 1), 1)
    ws.column_dimensions["A"].width = MAX_COL_WIDTH


def filter_frame(df: pd.DataFrame, campaign: str, level: int) -> pd.DataFrame:
    if df.empty:
        return df
    mask = pd.Series(True, index=df.index)
    if "campaign" in df.columns:
        mask &= df["campaign"].astype(str) == campaign
    if "kem_level" in df.columns:
        numeric_kem = pd.to_numeric(df["kem_level"], errors="coerce")
        mask &= numeric_kem == level
    return df[mask]


def build_report(
    campaign: str,
    level: int,
    frames: dict[str, pd.DataFrame],
    console_lines: list[str],
    output_dir: Path,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for key in ("summary", "all_runs", "failed_runs", "log_manifest", "validation_issues"):
        write_dataframe_sheet(
            workbook,
            SHEET_TITLES[key],
            filter_frame(frames[key], campaign, level),
        )

    study_prefix = "A-kem" if campaign == "A_B1_VS_B2" else "B-kem"
    level_tokens = (f"{study_prefix}{level}-", f"kem={level}")
    selected_console = [
        line for line in console_lines
        if any(token in line for token in level_tokens)
    ]
    write_console_sheet(workbook, SHEET_TITLES["console_log"], selected_console)

    campaign_dir = output_dir / campaign
    campaign_dir.mkdir(parents=True, exist_ok=True)
    out_path = campaign_dir / f"ML-KEM-{level}_rapport.xlsx"
    workbook.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", type=Path, default=Path("."))
    parser.add_argument("--output-dir", type=Path, default=Path("workbooks"))
    args = parser.parse_args()

    input_dir = args.input_dir
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Lecture des fichiers depuis : {input_dir}")
    frames: dict[str, pd.DataFrame] = {}
    for key, filename in INPUT_FILES.items():
        frames[key] = normalize_frame(load_csv(input_dir / filename))

    console_lines = load_console_logs(input_dir)

    campaigns_found: set[str] = set()
    levels_found: set[int] = set()
    for df in frames.values():
        if df.empty:
            continue
        if "campaign" in df.columns:
            campaigns_found.update(
                str(value) for value in df["campaign"].dropna().unique()
                if str(value) in KNOWN_CAMPAIGNS
            )
        if "kem_level" in df.columns:
            numeric = pd.to_numeric(df["kem_level"], errors="coerce").dropna()
            levels_found.update(int(value) for value in numeric.unique())

    campaigns = [c for c in KNOWN_CAMPAIGNS if c in campaigns_found] or KNOWN_CAMPAIGNS
    levels = sorted(levels_found) if levels_found else KNOWN_SECURITY_LEVELS

    print(f"Campagnes détectées : {campaigns}")
    print(f"Niveaux de sécurité détectés : {levels}")

    written: list[Path] = []
    for campaign in campaigns:
        for level in levels:
            all_runs = filter_frame(frames["all_runs"], campaign, level)
            if all_runs.empty:
                continue
            out_path = build_report(campaign, level, frames, console_lines, output_dir)
            print(f"  -> {campaign}/{out_path.name} écrit ({len(all_runs)} runs)")
            written.append(out_path)

    print(f"\nTerminé : {len(written)} classeur(s) Excel généré(s) dans {output_dir}")


if __name__ == "__main__":
    main()
